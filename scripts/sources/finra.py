"""FINRA Margin Statistics - Wertpapierkredite von US-Privatanlegern.

FINRA bietet keine Schnittstelle, nur eine XLSX-Datei auf der Themenseite. Der
Dateipfad ist stabil, aber nicht garantiert; deshalb wird bei einem 404 die
Themenseite nach dem aktuellen Link durchsucht, statt sofort aufzugeben.
"""

from __future__ import annotations

import io
import re

from openpyxl import load_workbook

from common import Reihe, hole, zahl

DIREKT = "https://www.finra.org/sites/default/files/2021-03/margin-statistics.xlsx"
THEMENSEITE = "https://www.finra.org/rules-guidance/key-topics/margin-accounts/margin-statistics"

MONATE = {m: i for i, m in enumerate(
    ["jan", "feb", "mar", "apr", "may", "jun",
     "jul", "aug", "sep", "oct", "nov", "dec"], start=1)}


def _datei() -> bytes:
    try:
        return hole(DIREKT, timeout=120).content
    except Exception:                                  # noqa: BLE001
        seite = hole(THEMENSEITE, timeout=60).text
        treffer = re.search(r'href="([^"]+\.xlsx)"', seite, re.I)
        if not treffer:
            raise RuntimeError(
                "FINRA-XLSX weder unter dem bekannten Pfad noch als Link auf "
                "der Themenseite gefunden."
            )
        pfad = treffer.group(1)
        url = pfad if pfad.startswith("http") else f"https://www.finra.org{pfad}"
        return hole(url, timeout=120).content


def _tag(zellwert) -> str | None:
    """FINRA mischt 'Jan-97', echte Datumswerte und '2026-07'."""
    if hasattr(zellwert, "year"):
        return f"{zellwert.year:04d}-{zellwert.month:02d}-01"
    text = str(zellwert or "").strip()
    treffer = re.match(r"^([A-Za-z]{3})[- ]?(\d{2,4})$", text)
    if treffer:
        monat = MONATE.get(treffer.group(1).lower())
        jahr = int(treffer.group(2))
        if monat:
            if jahr < 100:
                jahr += 1900 if jahr >= 70 else 2000
            return f"{jahr:04d}-{monat:02d}-01"
    treffer = re.match(r"^(\d{4})[-/](\d{1,2})", text)
    if treffer:
        return f"{int(treffer.group(1)):04d}-{int(treffer.group(2)):02d}-01"
    return None


def abrufen(key: str, args: dict) -> Reihe:
    blatt = load_workbook(io.BytesIO(_datei()), data_only=True).worksheets[0]
    zeilen = list(blatt.iter_rows(values_only=True))

    # Kopfzeile ueber den Begriff "debit" suchen - die Spaltenbeschriftung
    # lautet je nach Jahrgang "Debit Balances in Customers' Securities Margin
    # Accounts" oder kuerzer.
    kopf_index = spalte_debit = None
    for i, zeile in enumerate(zeilen[:20]):
        for j, feld in enumerate(zeile):
            if feld and "debit" in str(feld).lower():
                kopf_index, spalte_debit = i, j
                break
        if kopf_index is not None:
            break

    if spalte_debit is None:
        raise RuntimeError(
            "Spalte mit den Debit Balances in der FINRA-Datei nicht gefunden - "
            "Layout geaendert, scripts/sources/finra.py anpassen."
        )

    reihe = Reihe(key=key, quelle="FINRA Margin Statistics")
    for zeile in zeilen[kopf_index + 1:]:
        if not zeile or spalte_debit >= len(zeile):
            continue
        tag = _tag(zeile[0])
        wert = zahl(zeile[spalte_debit])
        if tag and wert is not None:
            reihe.punkte.append((tag, wert))
    return reihe
